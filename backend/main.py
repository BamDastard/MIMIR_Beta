from fastapi import FastAPI, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional, List, Dict, Any
from backend.core.ai import mimir_ai
from backend.core.memory import mimir_memory
from backend.core.voice import mimir_voice
from backend.core.daily_journal import daily_journal
from backend.core.news import news_manager
import uvicorn
import base64
from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook
from PIL import Image
import io
import google.generativeai as genai
import os
import json
from dotenv import load_dotenv
from google.oauth2 import id_token
from google.auth.transport import requests as google_requests

load_dotenv()
GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY")
GOOGLE_CLIENT_ID = os.getenv("GOOGLE_CLIENT_ID") # Add this to env
print(f"Backend initialized with GOOGLE_CLIENT_ID: {GOOGLE_CLIENT_ID[:5]}..." if GOOGLE_CLIENT_ID else "WARNING: GOOGLE_CLIENT_ID is missing")
genai.configure(api_key=GOOGLE_API_KEY)

from fastapi.staticfiles import StaticFiles

app = FastAPI(title="MIMIR API")

# Allow CORS for local frontend and Cloud Run
# In production, replace "*" with specific Cloud Run URL
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], # Allow all for now to ensure connectivity
    allow_origin_regex=r"https://.*\.run\.app", # Explicitly allow Cloud Run subdomains
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

from backend.core.user_manager import user_manager

# Authentication Middleware
from fastapi import Request, HTTPException, status, Response

async def verify_google_token(token: str):
    try:
        # Specify the CLIENT_ID of the app that accesses the backend:
        id_info = id_token.verify_oauth2_token(token, google_requests.Request(), GOOGLE_CLIENT_ID)
        return id_info
    except ValueError as e:
        print(f"Token verification error: {e}")
        return None

@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    if request.method == "OPTIONS":
        return await call_next(request)
        
    # Skip auth for health check and docs
    if request.url.path in ["/", "/docs", "/openapi.json"]:
        return await call_next(request)

    # Allow unauthenticated access to user check endpoints (to prevent catch-22)
    # Actually, /user/me and /user/onboard REQUIRE auth token, but might not have a profile yet.
    
    auth_header = request.headers.get('Authorization')
    
    # Development Mode (No Client ID) - Permissive
    if not GOOGLE_CLIENT_ID:
        # Mock user for dev
        request.state.user_auth_id = "dev_user_123"
        request.state.user_email = "dev@example.com"
        request.state.user_name = "Dev User"
        return await call_next(request)

    # Production Mode - Strict
    if not auth_header or not auth_header.startswith('Bearer '):
        return Response(status_code=401, content=json.dumps({"error": "Missing or invalid token"}), media_type="application/json")

    token = auth_header.split(' ')[1]
    user_info = await verify_google_token(token)
    
    if not user_info:
        return Response(status_code=401, content=json.dumps({"error": "Invalid token"}), media_type="application/json")
    
    # Inject User Info into Request State
    request.state.user_auth_id = user_info['sub']
    request.state.user_email = user_info.get('email', '')
    request.state.user_name = user_info.get('name', 'User')
    
    # Note: We do NOT check for profile existence here. 
    # That is handled by the endpoints. /user/onboard needs to work even if no profile exists.

    response = await call_next(request)
    return response

# Mount journal attachments
os.makedirs("journal_attachments", exist_ok=True)
app.mount("/journal_attachments", StaticFiles(directory="journal_attachments"), name="journal_attachments")

class ChatRequest(BaseModel):
    message: str
    personality_intensity: int = 75 # 0-100, default 75%
    mute: bool = False # If true, skip audio generation

class ChatResponse(BaseModel):
    text: str
    audio_base64: Optional[str] = None
    tools_used: List[str] = []
    tool_results: List[Dict[str, Any]] = []

@app.get("/")
def read_root():
    return {"status": "MIMIR is awake"}

# User Management Endpoints
@app.get("/user/me")
def get_current_user(request: Request):
    auth_id = request.state.user_auth_id
    profile = user_manager.get_profile(auth_id)
    if not profile:
        # Auto-onboard if profile missing (ephemeral storage fix)
        email = request.state.user_email
        name = request.state.user_name
        profile = user_manager.create_profile(auth_id, email, name)
    return profile

class OnboardRequest(BaseModel):
    display_name: str

@app.post("/user/onboard")
def onboard_user(request: Request, data: OnboardRequest):
    auth_id = request.state.user_auth_id
    email = request.state.user_email
    
    if not data.display_name.strip():
        raise HTTPException(status_code=400, detail="Display name is required")
        
    profile = user_manager.create_profile(auth_id, email, data.display_name)
    return profile

from fastapi.responses import StreamingResponse
import asyncio

# Ensure necessary directories exist on startup
# Ensure necessary directories exist on startup
@app.on_event("startup")
async def startup_event():
    dirs = ["daily_logs", "journal_attachments", "journal_entries", "calendars", "temp", "mimir_memory_db"]
    for d in dirs:
        os.makedirs(d, exist_ok=True)
        print(f"[STARTUP] Ensured directory exists: {d}")

    # Ensure user_profiles.json exists
    if not os.path.exists("user_profiles.json"):
        with open("user_profiles.json", "w") as f:
            json.dump({}, f)
        print("[STARTUP] Created empty user_profiles.json")

    # Run maintenance tasks
    try:
        daily_journal.cleanup_old_logs()
    except Exception as e:
        print(f"[STARTUP] Log cleanup failed: {e}")

@app.post("/chat")
async def chat(request: Request, body: ChatRequest):
    try:
        from datetime import datetime
        user_msg = body.message
        personality = body.personality_intensity
        
        # Get Authenticated User
        auth_id = request.state.user_auth_id
        profile = user_manager.get_profile(auth_id)
        
        if not profile:
            raise HTTPException(status_code=403, detail="User not onboarded")
            
        user_id = auth_id # Use Auth ID for internal storage
        display_name = profile.display_name # Use Display Name for AI Context
        
        async def event_generator():
            try:
                # 1. Recall Context for specific user
                daily_journal.log_interaction(user_id, "chat", f"User: {user_msg}")
                context = mimir_memory.recall(user_msg, user_id=user_id)
                
                # 2. Add current date/time to context
                current_time = datetime.now().strftime("%A, %B %d, %Y at %I:%M %p")
                time_context = f"Current Date and Time: {current_time}\\nUser Name: {display_name}"
                
                if context:
                    context = f"{time_context}\\n\\n{context}"
                else:
                    context = time_context
    
                # 2.5 Check for Daily Journal Triggers
                await daily_journal.check_end_of_day(user_id)
                
                if daily_journal.check_prompt_needed(user_id):
                    daily_journal.mark_prompted(user_id)
                    context += "\\n\\n[SYSTEM NOTE: It is after 7:00 PM and the user has not recorded much today. Gently ask them how their day went and if they have anything to add to their daily log.]"
                
                # 3. Generate Response (Parallel Audio via Queue with Ordered Collection)
                output_queue = asyncio.Queue()
                audio_task_queue = asyncio.Queue()
                
                async def generate_audio_task(text):
                    """Helper to generate audio (runs in background)"""
                    try:
                        return await asyncio.to_thread(mimir_voice.speak, text)
                    except Exception as e:
                        print(f"Audio generation failed: {e}")
                        return None

                async def audio_collector():
                    """Consumes audio tasks in order and sends results to output"""
                    while True:
                        task = await audio_task_queue.get()
                        if task is None:
                            break
                        
                        try:
                            # Await the task to ensure order
                            wav_bytes = await task
                            if wav_bytes:
                                audio_b64 = base64.b64encode(wav_bytes).decode('utf-8')
                                await output_queue.put(json.dumps({
                                    "type": "audio_chunk",
                                    "audio_base64": audio_b64
                                }) + "\n")
                        except Exception as e:
                            print(f"Audio collection failed: {e}")

                async def text_processor():
                    try:
                        text_buffer = ""
                        import re

                        # Pass the system prompt as user_input
                        google_token = request.headers.get("X-Google-Access-Token")
                        async for event in mimir_ai.generate_response_stream(user_msg, context, personality_intensity=personality, user_id=user_id, google_token=google_token):
                            # Pass through all events to frontend immediately
                            await output_queue.put(json.dumps(event) + "\n")

                            if event["type"] == "response_chunk":
                                chunk_text = event["text"]
                                text_buffer += chunk_text
                                
                                if not body.mute:
                                    # Split by sentence endings (. ? ! followed by space or newline)
                                    parts = re.split(r'(?<=[.!?])\s+', text_buffer)
                                    if len(parts) > 1:
                                        text_buffer = parts.pop() # Keep last part
                                        for part in parts:
                                            clean_text = re.sub(r'\[TOOL:.*?\]', '', part, flags=re.DOTALL).strip()
                                            if clean_text:
                                                # Create task and queue it to preserve order
                                                task = asyncio.create_task(generate_audio_task(clean_text))
                                                await audio_task_queue.put(task)
                                            
                            elif event["type"] == "tool_call":
                                # Flush buffer on tool call
                                if not body.mute and text_buffer.strip():
                                    clean_text = re.sub(r'\[TOOL:.*?\]', '', text_buffer, flags=re.DOTALL).strip()
                                    if clean_text:
                                        task = asyncio.create_task(generate_audio_task(clean_text))
                                        await audio_task_queue.put(task)
                                    text_buffer = ""

                            elif event["type"] == "response":
                                response_text = event["text"]
                                tools_used = event["tools_used"]
                                tool_results = event["tool_results"]
                                
                                # Flush remaining buffer
                                if not body.mute and text_buffer.strip():
                                    clean_text = re.sub(r'\[TOOL:.*?\]', '', text_buffer, flags=re.DOTALL).strip()
                                    if clean_text:
                                        task = asyncio.create_task(generate_audio_task(clean_text))
                                        await audio_task_queue.put(task)
                                    text_buffer = ""
                                
                                # 3. Remember Interaction
                                mimir_memory.remember(f"User: {user_msg}\\nMIMIR: {response_text}", user_id=user_id)
                                daily_journal.log_interaction(user_id, "chat", f"MIMIR: {response_text}")
                                if tools_used:
                                    daily_journal.log_interaction(user_id, "tool_use", {"tools": tools_used, "results": tool_results})

                    except Exception as e:
                        print(f"Error in text_processor: {e}")
                        import traceback
                        traceback.print_exc()
                        await output_queue.put(json.dumps({"type": "error", "content": "An internal error occurred."}) + "\n")
                    finally:
                        # Signal end of audio tasks
                        await audio_task_queue.put(None)

                # Run processor and collector concurrently
                async def coordinator():
                    await asyncio.gather(text_processor(), audio_collector())
                    await output_queue.put(None) # Signal end of stream

                asyncio.create_task(coordinator())

                # Consumer loop
                while True:
                    item = await output_queue.get()
                    if item is None:
                        break
                    yield item

            except Exception as e:
                print(f"Error in event generator: {e}")
                import traceback
                traceback.print_exc()
                yield json.dumps({"type": "error", "content": "An internal error occurred."}) + "\n"

        return StreamingResponse(event_generator(), media_type="application/x-ndjson")
    except Exception as e:
        print(f"Error in chat endpoint: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

from backend.core.planning import plan_day

@app.post("/chat/plan_day")
async def chat_plan_day(request: Request, body: ChatRequest):
    try:
        from datetime import datetime
        auth_id = request.state.user_auth_id
        profile = user_manager.get_profile(auth_id)
        
        if not profile:
            raise HTTPException(status_code=403, detail="User not onboarded")
            
        user_id = auth_id
        display_name = profile.display_name
        
        # 1. Generate Plan Data
        google_token = request.headers.get("X-Google-Access-Token")
        plan_data = await plan_day(user_id, google_token=google_token)
        user_msg = plan_data["system_prompt"]
        personality = body.personality_intensity
        
        async def event_generator():
            try:
                # 2. Log Interaction
                daily_journal.log_interaction(user_id, "action", "Started daily planning session")
                
                # 3. Generate Response (Reuse logic from /chat)
                # We don't need memory recall for this specific system prompt, 
                # but we might want to pass the summaries as context if we want to be cleaner.
                # For now, the system prompt contains everything.
                
                context = f"Current Date and Time: {datetime.now().strftime('%A, %B %d, %Y at %I:%M %p')}\\nUser Name: {display_name}"
                
                output_queue = asyncio.Queue()
                audio_task_queue = asyncio.Queue()
                
                # ... Reuse audio/text processing logic ...
                # Since the logic is identical to /chat, we should ideally refactor.
                # But for now, I will duplicate the coordinator/processor logic to ensure it works without breaking /chat.
                
                async def generate_audio_task(text):
                    try:
                        return await asyncio.to_thread(mimir_voice.speak, text)
                    except Exception as e:
                        print(f"Audio generation failed: {e}")
                        return None

                async def audio_collector():
                    while True:
                        task = await audio_task_queue.get()
                        if task is None: break
                        try:
                            wav_bytes = await task
                            if wav_bytes:
                                audio_b64 = base64.b64encode(wav_bytes).decode('utf-8')
                                await output_queue.put(json.dumps({
                                    "type": "audio_chunk",
                                    "audio_base64": audio_b64
                                }) + "\n")
                        except Exception as e:
                            print(f"Audio collection failed: {e}")

                async def text_processor():
                    try:
                        text_buffer = ""
                        import re

                        # Pass the system prompt as user_input
                        google_token = request.headers.get("X-Google-Access-Token")
                        async for event in mimir_ai.generate_response_stream(user_msg, context, personality_intensity=personality, user_id=user_id, google_token=google_token):
                            await output_queue.put(json.dumps(event) + "\n")

                            if event["type"] == "response_chunk":
                                chunk_text = event["text"]
                                text_buffer += chunk_text
                                if not body.mute:
                                    parts = re.split(r'(?<=[.!?])\s+', text_buffer)
                                    if len(parts) > 1:
                                        text_buffer = parts.pop()
                                        for part in parts:
                                            clean_text = re.sub(r'\[TOOL:.*?\]', '', part, flags=re.DOTALL).strip()
                                            if clean_text:
                                                task = asyncio.create_task(generate_audio_task(clean_text))
                                                await audio_task_queue.put(task)
                            
                            elif event["type"] == "tool_call":
                                if not body.mute and text_buffer.strip():
                                    clean_text = re.sub(r'\[TOOL:.*?\]', '', text_buffer, flags=re.DOTALL).strip()
                                    if clean_text:
                                        task = asyncio.create_task(generate_audio_task(clean_text))
                                        await audio_task_queue.put(task)
                                    text_buffer = ""

                            elif event["type"] == "response":
                                response_text = event["text"]
                                if not body.mute and text_buffer.strip():
                                    clean_text = re.sub(r'\[TOOL:.*?\]', '', text_buffer, flags=re.DOTALL).strip()
                                    if clean_text:
                                        task = asyncio.create_task(generate_audio_task(clean_text))
                                        await audio_task_queue.put(task)
                                    text_buffer = ""
                                
                                # Remember this interaction
                                mimir_memory.remember(f"MIMIR (Daily Plan): {response_text}", user_id=user_id)

                    except Exception as e:
                        print(f"Error in text_processor: {e}")
                        await output_queue.put(json.dumps({"type": "error", "content": "An internal error occurred."}) + "\n")
                    finally:
                        await audio_task_queue.put(None)

                async def coordinator():
                    await asyncio.gather(text_processor(), audio_collector())
                    await output_queue.put(None)

                asyncio.create_task(coordinator())

                while True:
                    item = await output_queue.get()
                    if item is None: break
                    yield item

            except Exception as e:
                print(f"Error in event generator: {e}")
                yield json.dumps({"type": "error", "content": "An internal error occurred."}) + "\n"

        return StreamingResponse(event_generator(), media_type="application/x-ndjson")
    except Exception as e:
        print(f"Error in plan_day endpoint: {e}")
        raise HTTPException(status_code=500, detail=str(e))

def read_document_content(file_path: str, content: bytes = None) -> str:
    """Reads content from a file path or bytes"""
    try:
        if content is None:
            with open(file_path, "rb") as f:
                content = f.read()
        
        text = ""
        filename = os.path.basename(file_path)
        
        if filename.endswith('.txt') or filename.endswith('.csv'):
            text = content.decode('utf-8')
            
        elif filename.endswith('.pdf'):
            pdf_reader = PdfReader(io.BytesIO(content))
            text = "\\n".join([page.extract_text() for page in pdf_reader.pages])
            
        elif filename.endswith(('.doc', '.docx')):
            doc = Document(io.BytesIO(content))
            text = "\\n".join([paragraph.text for paragraph in doc.paragraphs])
            
        elif filename.endswith(('.xls', '.xlsx')):
            wb = load_workbook(io.BytesIO(content))
            text_parts = []
            for sheet in wb.worksheets:
                text_parts.append(f"Sheet: {sheet.title}")
                for row in sheet.iter_rows(values_only=True):
                    text_parts.append(" | ".join([str(cell) if cell is not None else "" for cell in row]))
            text = "\\n".join(text_parts)
            
        elif filename.endswith(('.jpg', '.jpeg', '.png', '.gif', '.bmp')):
            # For direct reading, we might return a description or handle it in AI core
            # But for memory storage, we use Vision
            model = genai.GenerativeModel('gemini-2.5-pro')
            image = Image.open(io.BytesIO(content))
            response = model.generate_content([
                "Describe this image in detail, including any text visible in the image:",
                image
            ])
            text = f"Image: {filename}\\n{response.text}"
            
        return text
    except Exception as e:
        print(f"Error reading document {file_path}: {e}")
        return ""

@app.post("/upload")
async def upload_document(request: Request, file: UploadFile = File(...)):
    """Upload a document to MIMIR's memory"""
    user_id = request.state.user_auth_id
    try:
        content = await file.read()
        text = read_document_content(file.filename, content)
        
        if not text.strip():
            return {"status": "error", "message": "No text content could be extracted from the file"}
        
        # Store in memory for specific user
        mimir_memory.remember(text, user_id=user_id, metadata={"source": file.filename, "type": "document"})
        
        return {"status": "success", "message": f"Document '{file.filename}' has been added to MIMIR's memory"}
    except Exception as e:
        return {"status": "error", "message": f"Error processing file: {str(e)}"}

@app.post("/upload_temp")
async def upload_temp(file: UploadFile = File(...)):
    """Upload a file to a temp directory and return the path"""
    try:
        temp_dir = os.path.join(os.getcwd(), "temp")
        os.makedirs(temp_dir, exist_ok=True)
        
        file_path = os.path.join(temp_dir, file.filename)
        with open(file_path, "wb") as f:
            content = await file.read()
            f.write(content)
            
        return {"path": file_path}
    except Exception as e:
        return {"error": str(e)}

# Calendar endpoints
from backend.core.calendar import CalendarManager

@app.get("/calendar/events")
async def get_calendar_events(request: Request, start_date: str = None, end_date: str = None):
    """Get all calendar events or filter by date range"""
    user_id = request.state.user_auth_id
    google_token = request.headers.get("X-Google-Access-Token")
    print(f"[DEBUG] get_calendar_events: Token present? {bool(google_token)}")
    calendar_manager = CalendarManager(user_id=user_id, google_token=google_token)
    
    # Trigger background sync if token is available
    if google_token:
        import threading
        print("[DEBUG] Triggering background sync_down")
        threading.Thread(target=calendar_manager.sync_down).start()
        
    events = calendar_manager.get_events(start_date, end_date)
    return {"events": events}

@app.post("/calendar/events")
async def create_calendar_event(request: Request, event: dict):
    """Create a new calendar event"""
    user_id = request.state.user_auth_id
    google_token = request.headers.get("X-Google-Access-Token")
    calendar_manager = CalendarManager(user_id=user_id, google_token=google_token)
    result = calendar_manager.create_event(
        subject=event['subject'],
        date=event['date'],
        start_time=event.get('start_time'),
        end_time=event.get('end_time'),
        details=event.get('details')
    )
    return result

@app.put("/calendar/events/{event_id}")
async def update_calendar_event(request: Request, event_id: str, event: dict):
    """Update a calendar event"""
    user_id = request.state.user_auth_id
    google_token = request.headers.get("X-Google-Access-Token")
    calendar_manager = CalendarManager(user_id=user_id, google_token=google_token)
    result = calendar_manager.update_event(event_id, **{k: v for k, v in event.items() if k != 'user_id'})
    return result

@app.delete("/calendar/events/{event_id}")
async def delete_calendar_event(request: Request, event_id: str):
    """Delete a calendar event"""
    user_id = request.state.user_auth_id
    google_token = request.headers.get("X-Google-Access-Token")
    calendar_manager = CalendarManager(user_id=user_id, google_token=google_token)
    success = calendar_manager.delete_event(event_id)
    return {"success": success}

@app.get("/news/top")
async def get_top_news(request: Request, refresh: bool = False):
    """Get news headlines, personalized if user has preferences."""
    try:
        # Try to get user ID from request state (set by auth middleware)
        # Note: This endpoint might be called without auth in some cases, so we handle that.
        auth_id = getattr(request.state, "user_auth_id", None)
        
        query = None
        if auth_id:
            preferences = user_manager.get_preferences(auth_id)
            if preferences:
                import random
                # Pick one random preference to keep the feed fresh and focused
                query = random.choice(preferences)
                print(f"[NEWS] Fetching news for preference: {query}")
        
        news_items = news_manager.get_news(query=query, force_refresh=refresh)
        
        # If preference search yielded no results, fallback to top news
        if not news_items and query:
             print(f"[NEWS] No results for '{query}', falling back to top news.")
             news_items = news_manager.get_top_news(force_refresh=refresh)
             
        return {"news": news_items}
    except Exception as e:
        print(f"[NEWS] Error in get_top_news: {e}")
        return {"news": news_manager.get_top_news(force_refresh=refresh)}

@app.post("/news/log")
async def log_news_access(request: Request, title: str = Form(...), url: str = Form(...)):
    """Log that a user accessed a news item."""
    user_id = request.state.user_auth_id
    daily_journal.log_interaction(user_id, "action", f"Read news: {title} ({url})")
    return {"status": "logged"}

@app.post("/open_file")
async def open_file(path: str = Form(...)):
    """Opens a file on the host system."""
    import subprocess
    import platform
    
    if not os.path.exists(path):
        return {"error": "File not found"}
        
    try:
        # Disable in Cloud Run (check for K_SERVICE env var)
        if os.getenv("K_SERVICE"):
             return {"error": "This feature is disabled in the cloud environment."}

        if platform.system() == 'Windows':
            os.startfile(path)
        elif platform.system() == 'Darwin':
            subprocess.call(('open', path))
        else:
            subprocess.call(('xdg-open', path))
        return {"status": "success"}
    except Exception as e:
        return {"error": str(e)}

@app.get("/journal/{date_str}")
async def get_journal_entry(request: Request, date_str: str):
    """Get the journal entry for a specific date."""
    user_id = request.state.user_auth_id
    safe_id = "".join([c for c in user_id if c.isalnum() or c in (' ', '_', '-')]).strip()
    MIMIR_DATA_DIR = os.getenv("MIMIR_DATA_DIR")
    if MIMIR_DATA_DIR:
        journal_dir = os.path.join(MIMIR_DATA_DIR, "journal_entries")
    else:
        journal_dir = os.path.join(os.getcwd(), "journal_entries")
    journal_path = os.path.join(journal_dir, f"{user_id}_{date_str}.json")
    
    if not os.path.exists(journal_path):
        return {"error": "Journal entry not found"}
        
    try:
        with open(journal_path, 'r') as f:
            return json.load(f)
    except Exception as e:
        return {"error": f"Failed to load journal: {str(e)}"}

if __name__ == "__main__":
    uvicorn.run("backend.main:app", host="0.0.0.0", port=8000, reload=True)
